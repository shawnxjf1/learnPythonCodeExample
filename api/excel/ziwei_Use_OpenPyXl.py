#encoding=UTF-8
'''
Created on 2016年9月4日
@author: shawn
'''
import openpyxl
'''
openpyxl 一些基本概念：
workbook 由 worksheets组成，worksheet包含columns rows，a  box at a particular row and columns is called a cell.
The grid of cell marks up a sheet.
'''

'''
问题：
print('row=s%,key=%s,value=%s' % (index,key,mapAver[key]))
ValueError: unsupported format character ',' (0x2c) at index 6
'''

'''
改进：1.异常不会捕捉，导致批量中一条出错整个程序抛出
      2.批量中  一定要通过异常捕捉到时哪一条引起的异常 打印对应的key值。
'''

'''
注意点：
openPyxl 只能操作xlsx后缀的excel文件不能操作xls后缀的
'''
##定义两个全局变量，
#excel  列原始值
mapValList={}
#excel  列平均值
mapAver={}

def initMap(sheet):
    for i in range(2, 52903):
        c = sheet.cell(row=i, column=1)
        print(type(c.value))  ##根据excel 中的值定如若是 2085.750.  type(c.value)就是str，2085.750 就是float
        # 这里的cell(1,1) 就是excel中的第一行第一列，行头和列头不算在内
        try:
            ##excel 中数据多种多样，所以在清洗数据时候  except 捕捉异常，以让剩余值进行下去，且打印有异常的数据
            fixedVal = cleanData(str(c.value))
            mapValList[fixedVal] = []
        except ValueError:
            print("initMap-valueError,%s" % (c.value))


#1. excel中 清理数据，先把第一列变成数值类型
def calZiweiExcel(fileName):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.active
    initMap(sheet)
    averageMap()
    writeMap(sheet)
    wb.save('ziwei.xlsx')

def averageMap():
    for key in mapValList.keys():
        list1 = mapValList[key]
        num = len(list1)
        sum_score = sum(list1)
        mapAver[key]=sum_score/num

def groupByKey(sheet):
    for i in range(2, 52903):
        c = sheet.cell(row=i, column=1)
        try:
            fixedVal = cleanData(str(c.value))
        except ValueError:
            print("valueError,%s" % (c.value))
        sheet.cell(row=i, column=4).value = fixedVal
        ## 错误： Type 'type' object is not subscriptable
        ## KeyError: 2,如果map中没有key 2 你去遍历会报错的，所以必须先填充
        valList = mapValList[fixedVal]

        ##不为空的话
        cellColumn2 = sheet.cell(row=i, column=2)
        print(cellColumn2.value)
        mapValList[fixedVal].append(cellColumn2.value)

def averageMap():
    for key in mapValList.keys():
        list1 = mapValList[key]
        try:
            num = len(list1)
            sum_score = sum(list1)
            mapAver[key] = sum_score / num
        except TypeError :
            print('typeError, the key=' + key)
        else:
            print('else ')


def writeMap(sheet):
    for index, key in enumerate(mapAver.keys()):
        try:
            #print('row=s%,key=%s,value=%s' % (index, key, mapAver[key]))
            sheet.cell(row=index, column=6).value = key
            sheet.cell(row=index, column=7).value = mapAver[key]
        except Exception:
            print('writeMap error,key='+ key)

##1.去掉前面的- 号，遇到第一个小数点就截取
'''excel 格式清理：比如传进来是1). 后面有个点  2085.750.   2). 2085.7 50. 中间有空格'''
def cleanData(numStr):
    if numStr[0] == '-':
        numStr = numStr[1:]
    if numStr.find('.')== 0:
        return numStr[0:numStr.index('.')]
    else:
        print('numstr not contain . ,numstr=s%' % (numStr))
    #fixeme
    return numStr








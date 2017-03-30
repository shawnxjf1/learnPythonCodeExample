#encoding=UTF-8
'''2016-09-26
@author: shawn
'''
import openpyxl
import datetime

#datetime __slot__
#assert name in ("utcoffset", "dst")

##输出：成功
#D:\同步盘\Project\拉卡拉行政\加班报销\(201604-加班交通费及餐费明细表.xlsx
def loadOverWorkExcel(fileName,beginIndex,endIndex,destFileName):
    wb = openpyxl.load_workbook(fileName)
    ## excel表中存放的数据
    sheet = wb.get_sheet_by_name("0412")
    ## 实际下班时间所在的列
    intOffTimeColumn = 3
    #sheet = wb.active

    ## 一个一个遍历excel的每一行
    for i in range(beginIndex,endIndex):
        c = sheet.cell(row=i, column=intOffTimeColumn)  ## 4
        '''注意 print语句 %s s在%后面'''
        print("c.value.type=%s and c.value=%s"%(type(c.value),c.value)) ##<class 'datetime.time'> and c.value=20:41:00 - (excel当中就是h:mm单元格的格式)

        ## 注意datetime 是属于datetime模块里的datetime类
        now = datetime.datetime.now()
        datatimeBaseLine = datetime.datetime(now.year,now.month,now.day,20,00,0,0)

        if c.value == None:
            print("row {0} is none".format(i))
            continue

        actualworkDateTime = patchActualWorkTime(c.value)
        week = getWeekFromTime(actualworkDateTime)
        shordateStr = getShortDateStr(actualworkDateTime)

        '''技术点： unsupported operand type(s) for -: 'datetime.time' and 'datetime.time' '''
        overTimeDelta = actualworkDateTime - datatimeBaseLine  ## unsupported operand type(s) for +: 'datetime.timedelta' and 'int'
        print("overTimeDelta=%s,and index=%s" % (overTimeDelta,i))

        ## 因为公司是半个小时计算一个加班时间的，所以这里要把时间转换成30小时（向下取整，即不到30 minute的去掉）。
        halfHourNumber = int(convertTimeDeltaTominutes(overTimeDelta)/30)/2 + 2
        cTimeCount = sheet.cell(row=i, column=12)
        cTimeCount.value = halfHourNumber

        cWeek = sheet.cell(row=i, column=13)
        cWeek.value = week

        cShortDate = sheet.cell(row=i, column=14)
        cShortDate.value = shordateStr

        #201604.xlsx
        wb.save(destFileName)

## 获取星期值
def getWeekFromTime(actualworkDateTime):
    return '星期%s ' % actualworkDateTime.strftime('%w')

## 获取日期值
def getShortDateStr(actualworkDateTime):
    return '%s月%s日 ' % (actualworkDateTime.strftime('%m'),actualworkDateTime.strftime('%d'))


## excel中的值会有很多种
## 1.当excel.cell中的值为：2016/5/9 20:18:52时候，则c.value.type=<class 'str'>。
## 2.当excel.cell中的值为 2016:5:9 20:18:52时候，则c.value.type=<class 'datetime'>
def patchActualWorkTime(actualWorkTime):

    # excel 中的值为：2016/5/9 20:18:52，则c.value.type=<class 'str'> and c.value=2016/5/4 20:04:49，所以需要改成20:18的格式
    ##判断类型想到了微信读书：python 86个注意点中说到判断的那一章
    now = datetime.datetime.now()
    if isinstance(actualWorkTime,datetime.datetime):
        actualworkDateTime = datetime.datetime(now.year, now.month, now.day, actualWorkTime.hour, actualWorkTime.minute,
                                   actualWorkTime.second, 0)

    # 2016/5/9 20:18:52，则c.value.type=<class 'str'>
    elif isinstance(actualWorkTime,str):
        actualWorkTime = convertStrToDateTime(actualWorkTime)
    print("actualworktime=%s" % actualWorkTime)
    return actualWorkTime

#抽取出时分秒组装成 datetime对象，2016/5/9 20:18:52 比如就会生成datetime对象为datetime.datetime(2016,5,9,20,18,52,0)
def convertStrToDateTime(timeStr):
    try:
        firstColonIndex = timeStr.index(':')
        firstSlashIndex = timeStr.index('/')
        hour = timeStr[firstColonIndex - 2:firstColonIndex]
        minute = timeStr[firstColonIndex + 1:firstColonIndex + 3]
        second = timeStr[firstColonIndex + 4:]

        year = timeStr[:firstSlashIndex]

        secondSlashIndex = timeStr.index('/', firstSlashIndex + 1)
        month = timeStr[firstSlashIndex + 1: secondSlashIndex]

        blankIndex = timeStr.index(' ', secondSlashIndex + 1)
        day = timeStr[secondSlashIndex + 1: blankIndex]

        year = year.strip()
        month = month.strip()
        day = day.strip()
        hour = hour.strip()
        second = second.strip()
    except TypeError:
        print("convertStrToDateTime-typeError,%s" % (timeStr))
    actualworkDateTime = datetime.datetime(int(year), int(month), int(day), int(hour), int(minute),int(second), 0)
    return actualworkDateTime;

## 转换timeDelta为秒
def convertTimeDeltaTominutes(overTimeDelta):
    '''
    # 注意timeDelta秒是转化之后的，比如overTimeDelat=0:41:00,and seconds = 2460（所有小时分钟折算成的秒）
    :param overTimeDelta:
    :return:
    '''
    seconds = overTimeDelta.seconds
    print('overTimeDelat=%s,and seconds = %s' % (overTimeDelta,seconds))
    # hour = overTimeDelta._seconds
    #minute = overTimeDelta._minutes
    totalMinute = seconds//60
    return totalMinute
    # totalMinute = hour * 60 + minute
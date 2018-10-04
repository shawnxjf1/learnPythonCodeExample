# encoding=UTF-8
'''2016-09-26
@author: shawn
'''
import openpyxl
import datetime
from  datetime import *
import os
import requests
# datetime __slot__
# assert name in ("utcoffset", "dst")
import re
import threading


def loadOverWorkExcel(fileName, beginIndex, endIndex, destFileName):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.get_sheet_by_name("baidu")
    # sheet = wb.active
    for i in range(beginIndex, endIndex):
        c = sheet.cell(row=i, column=4)
        '''注意 print语句 %s s在%后面'''
        print("c.value.type=%s and c.value=%s ,rowindex%s" % (
        type(c.value), c.value,i))  ##<class 'datetime.time'> and c.value=20:41:00 - (excel当中就是h:mm单元格的格式)

        cWeek = sheet.cell(row=i, column=5)
        try:
            cWeek.value = getProviderCity(c.value)
        except Exception as e:
            print('except:', e)
        # 201604.xlsx
        wb.save(destFileName)
        
## 跑一页也只跑2000多条
def getPhoneLocationFromExcelForBaiduWithMultiThread(fileName, beginIndex, endIndex, destFileName):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.get_sheet_by_name("baidu")
    threads = []
    sliceLenth = 1000;
    sliceBegin = 2
    sliceEnd = sliceBegin + sliceLenth
    # sheet = wb.active
    threadNum = int((endIndex - beginIndex)/1000)
    for i in range(1,threadNum):
        print("thread num i = %s" % (i))
        temp_thread = threading.Thread(target=dealExcelShell, args=(sheet,sliceBegin,sliceEnd,mexSize))
        threads.append(temp_thread)
        sliceBegin = i*1000 + 2
        sliceEnd = (i+1)*1000 + 2

    for t in threads:
        t.start()

    for tj in threads:
        tj.join()

    wb.save(destFileName)
    print ("Finished!")

def dealExcelShell(sheet,beginIndex,endIndex,maxsize):
    for i in range(beginIndex, endIndex):
        if (i >maxsize):
            print("begin then 30590 so return")
            return
        c = sheet.cell(row=i, column=4)
        '''注意 print语句 %s s在%后面'''
        print("c.value.type=%s and c.value=%s ,rowindex%s" % (
        type(c.value), c.value,i))  ##<class 'datetime.time'> and c.value=20:41:00 - (excel当中就是h:mm单元格的格式)
        if (type(c.value) != int):
            print("returen because type is not int,c.value.type=%s,rowIndex%s" % (type(c.value),i))


        cWeek = sheet.cell(row=i, column=5)
        try:
            cWeek.value = getProviderCity(c.value)
        except Exception as e:
            print('except:', e)
        # 201604.xlsx

def getPhoneLocationFromExcelForDidi(fileName, beginIndex, endIndex, destFileName):
    wb = openpyxl.load_workbook(fileName)
    sheet = wb.get_sheet_by_name("didi")
    # sheet = wb.active
    for i in range(beginIndex, endIndex):
        c = sheet.cell(row=i, column=4)
        '''注意 print语句 %s s在%后面'''
        print("c.value.type=%s and c.value=%s,rowIndex%s" % (
        type(c.value), c.value,i))  ##<class 'datetime.time'> and c.value=20:41:00 - (excel当中就是h:mm单元格的格式)
        if (type(c.value) != int):
            print("returen because type is not int,c.value.type=%s,rowIndex%s" % (type(c.value),i))

        cWeek = sheet.cell(row=i, column=5)
        try:
            cWeek.value = getProviderCity(c.value)
        except Exception as e:
            print('except:', e)
        # 201604.xlsx
        wb.save(destFileName)

def parseString(src, result):
    pat = []
    pat.append('(?<=归属地：</span>).+(?=<br />)')
    pat.append('(?<=卡类型：</span>).+(?=<br />)')
    pat.append('(?<=运营商：</span>).+(?=<br />)')
    pat.append('(?<=区号：</span>)\d+(?=<br />)')
    pat.append('(?<=邮编：</span>)\d+(?=<br />)')

    item = []
    for i in range(len(pat)):
        m = re.search(pat[i], src)
        if m:
            v = m.group(0)
            item.append(v)
    return item


def getPageCode(url):
    r = requests.get(url)
    ## can't use a string pattern on a bytes-like object'
    ## 查找了一下,是说3.0现在的参数更改了,现在读取的是bytes-like的,但参数要求是chart-like的,找了一下,加了个编码:
    data = r.content.decode('UTF-8')
    return data


def getProvider(phoneNum, result):
    url = "http://www.sjgsd.com/n/?q=%s" % phoneNum
    text = getPageCode(url)
    item = parseString(text, result)
    print("item is %s finished" % item)
    result.append((phoneNum, item))


def getProviderCity(phoneNum):
    url = "http://www.sjgsd.com/n/?q=%s" % phoneNum
    curDir = "/Users/shawn/eclipse_workspace_pdev_group/python_works/pythonLearnCodeExample/phones";
    fileName =  curDir + "/" + str(phoneNum) + ".html"
    text = "";
    if os.path.isfile(fileName):
        print("page fileName:%s exist." % fileName)
        text = readPage(fileName)
    else:
        print("page fileName:%s not exist..." % fileName)
        text = getPageCode(url)
        writePage(text,fileName)
    result = []
    item = parseString(text, result)
    print("item is %s finished" % item)
    return item[0]

def writePage(content,fullName):
    f = open(fullName, 'w')
    f.write(content)
    f.close()

def readPage(fileName):
    f = open(fileName,'r', encoding='UTF-8')  ##一定要加上encoding=utf-8 才能定位到某个参数，不然它会按顺序解析的
    fileContent = f.read()
    return fileContent

# write result to fileitem
def writeResult(result):
    f = open("result.log", "w")
    for num, item in result:
        f.write("%s:\t" % num)
        for i in item:
            f.write("%s,\t" % i)
        f.write("\n")
    f.close()


if __name__ == "__main__":
    item = getProviderCity("13904446048")
    print("the result is %s" % item)
    result = []
    for line in open("test.txt", "r"):
        phoneNum = line.strip(" \t\r\n")
        getProvider(phoneNum, result)
        print("%s is finished" % phoneNum)
    writeResult(result)
